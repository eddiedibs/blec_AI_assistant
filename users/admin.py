import os

from django.contrib import admin
from django.conf import settings
from users.models import Doctor, Patient, Appointment, PatientParent
from image_cropping import ImageCroppingMixin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import Group
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Field
from django.core.management import call_command
from django.urls import path
from datetime import datetime
import glob

from .views import backup_restore_view  # Import the custom view
from .models import CustomUser



# Define Backup Action
def backup_db_action(modeladmin, request, queryset):
    try:
        # Generate a backup file path (you can customize this part)
        backup_file_path = os.path.join(settings.BACKUP_DIR, f"db_backup_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.sql")
        
        # Call the custom backup command with the backup file path
        call_command('backup_db', backup_file=backup_file_path)
        modeladmin.message_user(request, f"Database backup started successfully! Backup file: {backup_file_path}", level="success")
    except Exception as e:
        modeladmin.message_user(request, f"Database backup failed: {str(e)}", level="error")

backup_db_action.short_description = "Realizar Backup de datos"


# Define Restore Action
def restore_db_action(modeladmin, request, queryset):
    try:
        # Use glob to find all files matching the db_backup_* pattern
        backup_files = glob.glob(os.path.join(settings.BACKUP_DIR, 'db_backup_*.sql'))

        if not backup_files:
            modeladmin.message_user(request, "No backup files found.", level="error")
            return

        # Get the most recent backup file (latest file)
        latest_backup_file = max(backup_files, key=os.path.getmtime)

        # Call the custom restore command with the latest backup file
        call_command('backup_db', restore=True, backup_file=latest_backup_file)
        
        modeladmin.message_user(request, f"Database restore started successfully from: {latest_backup_file}", level="success")
    except Exception as e:
        modeladmin.message_user(request, f"Database restore failed: {str(e)}", level="error")

restore_db_action.short_description = "Realizar restauración de datos"



def export_to_xlsx(modeladmin, request, queryset):
    model = modeladmin.model  # Get the model dynamically from the admin
    model_fields = [field.name for field in model._meta.get_fields() if isinstance(field, Field)]

    # Create an in-memory workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"

    # Add the header row (fields) horizontally
    for col_num, field_name in enumerate(model_fields, start=1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = field_name.capitalize()  # Set header as field name

    # Add the data rows
    for row_num, obj in enumerate(queryset, start=2):  # Start from row 2 (below headers)
        for col_num, field_name in enumerate(model_fields, start=1):
            value = getattr(obj, field_name)
            if value is None:
                value = "N/A"
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = str(value)

    # Create a response object that will contain the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    # Save the workbook to the response object
    wb.save(response)
    return response

export_to_xlsx.short_description = "Exportar reporte"



@admin.register(CustomUser)  # Registers the model in the admin panel
class CustomUserAdmin(UserAdmin):
    list_display = ['username', 'email', 'first_name', 'last_name', 'is_staff', 'is_doctor']
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering = ['username']
    list_filter = ['is_staff', 'is_superuser', 'is_doctor']
    actions = [export_to_xlsx, backup_db_action, restore_db_action]

    fieldsets = (
        ("Personal Information", {'fields': ('username', 'password', 'first_name', 'last_name', 'email')}),
        ("Permissions", {'fields': ('is_active', 'is_staff', 'is_superuser', 'is_doctor', 'groups', 'user_permissions')}),
        ("Important Dates", {'fields': ('last_login', 'date_joined')}),
    )

    add_fieldsets = (
        ("Create New User", {
            'classes': ('wide',),
            'fields': ('username', 'email', 'password1', 'password2', 'is_doctor'),
        }),
    )
admin.site.unregister(Group)

# Move CustomUser to "Autenticación y autorización"
admin.site._registry[CustomUser].app_label = "auth"

# Re-register Group after CustomUser so it's below it
admin.site.register(Group)

@admin.register(Doctor)
class DoctorAdmin(ImageCroppingMixin, admin.ModelAdmin):
    exclude = ('drive_url',)
    list_display = ['name', 'specialization', 'contact', 'email']
    search_fields = ['name', 'specialization']
    actions = [export_to_xlsx, backup_db_action, restore_db_action]

    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True  # Superadmins can modify anything

        if obj and request.user.is_doctor:
            return obj.user == request.user  # Only assigned doctor can modify

        return False  # Other users cannot modify


@admin.register(Patient)
class PatientAdmin(admin.ModelAdmin):
    list_display = ['name', 'get_age', 'get_parent_name']
    search_fields = ['name', 'parent__name']
    actions = [export_to_xlsx, backup_db_action, restore_db_action]

    def get_age(self, obj):
        return obj.age  # Calls the @property method

    def get_parent_name(self, obj):
        return obj.parent_name  

    get_age.short_description = "Edad"  # Set column title in admin panel
    get_parent_name.short_description = "Representante" 


@admin.register(PatientParent)
class PatientParentAdmin(admin.ModelAdmin):
    list_display = ['name', 'id_number', 'contact_phone_number', 'email']
    search_fields = ['name']
    actions = [export_to_xlsx, backup_db_action, restore_db_action]

@admin.register(Appointment)
class AppointmentAdmin(admin.ModelAdmin):
    list_display = ['doctor', 'patient', 'appointment_date', 'status']
    list_filter = ['appointment_date', 'doctor', 'status']
    search_fields = ['doctor__name', 'patient__name']
    actions = [export_to_xlsx, backup_db_action, restore_db_action]

    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True  # Superadmins can modify anything

        if obj and request.user.is_doctor:
            return obj.doctor.user == request.user  # Only assigned doctor can modify

        return False  # Other users cannot modify

    def has_delete_permission(self, request, obj=None):
        return self.has_change_permission(request, obj)  # Same logic for deletion

    # def has_view_permission(self, request, obj=None):
    #     return self.has_change_permission(request, obj)  # Same logic for deletion

